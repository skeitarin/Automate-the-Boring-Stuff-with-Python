#! python3
import os
from datetime import datetime

# TODO
# 1:エクスプローラでExcelファイルを選択
# 2:Excelファイルの中身を操作
# 3:Excelの中身でRedmineにチケットを作成

file_full_path = ''
user_id = ''
password = ''
ticket_info = [] # カンマ区切りで格納

# 1:エクスプローラでExcelファイルを選択
# ref:http://spcx8.hatenablog.com/entry/2017/12/24/112528
def SelectFile(file_types):
    import tkinter as tk
    import tkinter.filedialog as tkdialog

    root  = tk.Tk()
    file_full_path = tkdialog.askopenfilename(filetypes=file_types,initialdir=os.getcwd())
    root.withdraw()
    print(file_full_path)
    return file_full_path

file_full_path = SelectFile([('Excel files','*.xlsx')])

# 2:Excelファイルの中身を操作
import openpyxl
wb = openpyxl.load_workbook(file_full_path)
sheet = wb.get_sheet_by_name('Sheet1')

user_id = sheet.cell(column=2, row=1).value
password = sheet.cell(column=4, row=1).value
print(user_id)
print(password)
row_num = 4
while sheet.cell(column=1, row=row_num).value is not None:
    tmp_list = []
    tmp_list.append(sheet.cell(column=1, row=row_num).value)
    tmp_list.append(sheet.cell(column=2, row=row_num).value)
    tmp_list.append(sheet.cell(column=3, row=row_num).value)
    ticket_info.append(','.join(tmp_list))
    row_num+=1
print(ticket_info)

# 3:Excelの中身でRedmineにチケットを作成
# pip install selenium
# https://chromedriver.storage.googleapis.com/index.html?path=2.40/
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
base_url = 'http://104.197.206.76/redmine/'
browser = webdriver.Chrome(executable_path=os.path.abspath("/Users/Keita/Desktop/src/python/退屈なことはPythonにやらせよう/Sample/chromedriver"))
# open browser
browser.get(base_url)

# login
elem_id = browser.find_element_by_id('username')
elem_id.send_keys(user_id)
elem_pass = browser.find_element_by_id('password')
elem_pass.send_keys(password)
elem_pass.send_keys(Keys.ENTER)
# create tickets
for info in ticket_info:
    browser.get(base_url + 'projects/test2/issues/new')
    infos = info.split(',')
    elem_title = browser.find_element_by_id('issue_subject')
    elem_title.send_keys(infos[0] + ' : ' + datetime.now().strftime("%Y/%m/%d %H:%M:%S"))
    elem_desc = browser.find_element_by_id('issue_description')
    elem_desc.send_keys(infos[1])
    elem_priority = browser.find_element_by_id('issue_priority_id')
    elem_priority.send_keys(infos[2])

    elem_create = browser.find_element_by_name('commit')
    elem_create.submit()

sleep(3) # ブラウザクローズのために少し待機
browser.close()