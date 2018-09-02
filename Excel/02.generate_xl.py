import openpyxl, os

xl = './Excel/example_copy.xlsx'

if os.path.isfile(xl):
    os.remove(xl)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Spam Bacon'
print(wb.get_sheet_names())

# シートの追加
wb.create_sheet()
print(wb.get_sheet_names())
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.get_sheet_names())

# セルへの書き込み
sheet = wb.get_sheet_by_name('Middle Sheet')
sheet['A1'] = 'Hello Worlrd!!!'
print(sheet['A1'].value)

wb.save(xl)
