import openpyxl,os

wb = openpyxl.load_workbook('./Excel/example.xlsx')
print(type(wb))

# Sheet名の取得
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet3')
print(type(sheet))
print(sheet.title)

# Cellの取得
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'].value)

b = sheet['B1']
print('行' + str(b.row) + ', 列' + b.column + 'は' + b.value)
print(b.coordinate)
print(sheet.cell(row=1 ,column=2).value)
print(sheet.max_row)
print(sheet.max_column)

# Rangeの取得
range = sheet['A1':'C3']
for row in range:
    for cell in row:
        print(cell.coordinate, cell.value)
    print(' -- End of Row -- ')